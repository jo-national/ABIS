"""Bygger en test-fixture der bevidst udfordrer parseren:
- forskellige arknavne ("ABIS 2026", "2025", "Liste 2024", "Vejledning" uden år)
- header i forskellige rækker, forskellige kolonnenavne og -rækkefølge
- ugyldige ISIN, tomme rækker, dubletter på tværs af år
"""
from openpyxl import Workbook

wb = Workbook()
wb.remove(wb.active)

vejl = wb.create_sheet("Vejledning")
vejl["A1"] = "Denne fane indeholder ingen data."

s26 = wb.create_sheet("ABIS 2026")
s26.append(["Liste over aktiebaserede investeringsselskaber"])
s26.append([])
s26.append(["Navn", "ISIN-kode", "LEI-kode", "CVR-nr./SE-nr.", "Land"])
rows26 = [
    ["iShares Core MSCI World UCITS ETF", "IE00B4L5Y983", "549300SBOOZR51TG3W64", None, "Irland"],
    ["Xtrackers MSCI World UCITS ETF 1C", "IE00BJ0KDQ92", None, None, "Irland"],
    ["Sparindex INDEX Globale Aktier Min Risiko Akk KL", "DK0060747947", None, "12345678", "Danmark"],
    ["Ugyldig Fond ApS", "IKKE-EN-ISIN", None, None, "Danmark"],
    ["Amundi Prime All Country World UCITS ETF", "IE0009HF1MK9", None, None, "Irland"],
    [None, None, None, None, None],
    ["  vanguard ftse all-world ucits etf  ", " ie00bk5bqt80 ", None, None, "Irland"],
]
for r in rows26: s26.append(r)

s25 = wb.create_sheet("2025")
s25.append(["ISIN", "Name", "Country"])  # anden rækkefølge og engelske navne
rows25 = [
    ["IE00B4L5Y983", "iShares Core MSCI World UCITS ETF", "Ireland"],
    ["IE00BJ0KDQ92", "Xtrackers MSCI World UCITS ETF 1C", "Ireland"],
    ["LU0908500753", "Amundi Stoxx Europe 600 UCITS ETF", "Luxembourg"],
]
for r in rows25: s25.append(r)

s24 = wb.create_sheet("Liste 2024")
s24.append([])
s24.append(["Navn", "ISIN-kode"])
s24.append(["iShares Core MSCI World UCITS ETF", "IE00B4L5Y983"])
s24.append(["Gammel Fond (udgået)", "LU0274208692"])

wb.save("tests/fixture_abis.xlsx")
print("fixture skrevet")
