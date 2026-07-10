"""
parse.py — læser den arkiverede ABIS-Excel og bygger data/abis.json.

Version 2. Skrevet efter inspektion af den rigtige fil
(juni-2026-abis-liste-2021-2026.xlsx). Ændringer i forhold til version 1:

1. Kolonnerne kortlægges på deres FAKTISKE overskrifter, ikke på gæt.
   Ændrer Skattestyrelsen overskrifterne, fejler parseren højlydt.
2. Teksten "[tom]" i en celle betyder tom. Den behandles som tom.
3. Indkomstår aflæses af kolonnen "Registrerede år/Registered" og ikke af,
   hvilket faneblad rækken står på. Kolonnen dækker også 2020, som ikke har
   sit eget faneblad. Bemærk: Skattestyrelsen skriver nogle steder "2025.2026"
   med punktum i stedet for komma, så årstal udtrækkes med mønstergenkendelse.
   Fanebladsmedlemskabet bruges som uafhængig kontrol, og enhver uenighed
   rapporteres.
4. Fonde uden ISIN ("Udstedt uden" i ISIN-feltet) medtages nu. De identificeres
   på CVR-nummer i stedet. Tidligere blev de smidt væk — det gav forkerte svar.
5. Fondsnavnet sammensættes af afdelingsnavn og eventuel andelsklasse, som er
   det, en bruger søger efter.
6. Samme ISIN kan optræde med flere navne. Alle navne bevares.
7. ISIN'ets kontrolciffer efterprøves. Fejler det, beholdes fonden, men den
   markeres — vi sletter ikke data, Skattestyrelsen har offentliggjort.

Kør med --verify for kun at udskrive rapporten uden at skrive abis.json.
"""

from __future__ import annotations

import json
import re
import sys
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
DATA = ROOT / "data"
RAW = DATA / "raw"
STATE_FILE = DATA / "state.json"
OUT_FILE = DATA / "abis.json"
REPORT_FILE = DATA / "parse_report.txt"

SHEET_YEAR_RE = re.compile(r"^(20\d{2})$")
YEAR_RE = re.compile(r"20\d{2}")
ISIN_RE = re.compile(r"^[A-Z]{2}[A-Z0-9]{9}[0-9]$")
UDEN_ISIN_RE = re.compile(r"udstedt\s+uden", re.IGNORECASE)

FORVENTEDE = {
    "hjemsted": "skattemæssigt hjemsted/tax residence",
    "isin": "isin-kode/-code",
    "andelsklasse": "navn andelsklasse/name shareclass",
    "lei": "lei-kode/-code",
    "cvr": "cvr/se/tin",
    "afdeling": "navn afdeling/name sub-fund",
    "tin": "tin",
    "navn": "navn/name",
    "registreret": "registrerede år/registered",
    "afregistreret": "ikke registrerede år/deregistered",
}


def tekst(v) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return None if s in ("", "[tom]") else s


def isin_kontrolciffer_ok(isin: str) -> bool:
    """Efterprøver ISIN'ets sidste ciffer (Luhn-algoritmen, ISO 6166)."""
    cifre = "".join(str(int(c, 36)) if c.isalpha() else c for c in isin)
    total = 0
    for i, c in enumerate(reversed(cifre)):
        d = int(c)
        if i % 2 == 1:
            d *= 2
            if d > 9:
                d -= 9
        total += d
    return total % 10 == 0


def kortlæg_kolonner(header: tuple) -> dict[str, int]:
    normaliseret = {
        (str(c).strip().lower() if c is not None else ""): i
        for i, c in enumerate(header)
    }
    mapping: dict[str, int] = {}
    mangler: list[str] = []
    for nøgle, overskrift in FORVENTEDE.items():
        if overskrift in normaliseret:
            mapping[nøgle] = normaliseret[overskrift]
        else:
            mangler.append(overskrift)
    if mangler:
        raise KeyError(
            "Overskrifter mangler eller er ændret hos Skattestyrelsen: "
            + ", ".join(repr(m) for m in mangler)
        )
    return mapping


ORD_RE = re.compile(r"[a-zæøå0-9]+")


def _ord(s: str) -> set[str]:
    return set(ORD_RE.findall(s.lower()))


def visningsnavn(afdeling, andelsklasse, navn):
    """ISIN identificerer en andelsklasse, men andelsklassenavnet er nogle gange
    kun et suffiks ('Global Quant, EUR W') og andre gange det fulde navn.
    Vi vælger det navn, der indeholder det andet — og sammensætter, hvis ingen
    af dem gør. Så bevares både udsteder og klasse, uden dobbeltskrivning."""
    afd, kls = afdeling, andelsklasse
    if afd and kls:
        oa, ok = _ord(afd), _ord(kls)
        if oa <= ok:
            return kls          # andelsklassen er den fulde tekst
        if ok <= oa:
            return afd          # afdelingsnavnet er den fulde tekst
        return f"{afd} — {kls}"  # de supplerer hinanden
    return afd or kls or navn


def slug(s: str) -> str:
    s = s.lower()
    for a, b in [("æ", "ae"), ("ø", "oe"), ("å", "aa"), ("á", "a"), ("é", "e"), ("ü", "u")]:
        s = s.replace(a, b)
    return re.sub(r"-+", "-", re.sub(r"[^a-z0-9]+", "-", s)).strip("-")[:60]


def main(verify_only: bool = False) -> None:
    if not STATE_FILE.exists():
        sys.exit("FEJL: data/state.json findes ikke. Kør fetch.py først.")
    state = json.loads(STATE_FILE.read_text(encoding="utf-8"))
    xlsx_path = RAW / state["archived_as"]
    if not xlsx_path.exists():
        sys.exit(f"FEJL: arkiveret fil mangler: {xlsx_path}")

    rap: list[str] = [
        f"Parse-rapport — genereret {datetime.now(timezone.utc).isoformat(timespec='seconds')}",
        f"Kildefil: {state['original_filename']} (SHA-256 {state['sha256'][:16]}…)",
        f"Hentet fra: {state['file_url']}",
        "",
    ]

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)

    fonde: dict[str, dict] = {}
    ark_år: dict[int, set[str]] = {}
    tomme_rækker = 0
    ugyldige_isin: set[str] = set()
    uden_cvr: list[str] = []
    problemer: list[str] = []

    for sheet in wb.worksheets:
        m = SHEET_YEAR_RE.match(sheet.title.strip())
        if not m:
            rap.append(f"Ark '{sheet.title}': ikke et årstal — springes over (forventet).")
            continue
        ark_årstal = int(m.group(1))
        rows = sheet.iter_rows(values_only=True)
        try:
            kol = kortlæg_kolonner(next(rows))
        except (StopIteration, KeyError) as e:
            problemer.append(f"Ark '{sheet.title}': {e}")
            continue

        antal = 0
        for row in rows:
            if all(v is None for v in row):
                tomme_rækker += 1
                continue

            def c(nøgle: str):
                i = kol[nøgle]
                return tekst(row[i]) if i < len(row) else None

            rå_isin = c("isin")
            afdeling, andelsklasse, navn = c("afdeling"), c("andelsklasse"), c("navn")
            cvr = c("cvr")
            år = {int(y) for y in YEAR_RE.findall(c("registreret") or "")}

            if rå_isin and ISIN_RE.match(rå_isin.upper()):
                fond_id = rå_isin.upper()
                type_ = "isin"
                if not isin_kontrolciffer_ok(fond_id):
                    ugyldige_isin.add(fond_id)
            elif rå_isin and UDEN_ISIN_RE.search(rå_isin):
                type_ = "uden_isin"
                if cvr:
                    fond_id = f"CVR-{cvr}"
                else:
                    # Udenlandske selskaber har intet CVR. De identificeres på navn,
                    # så de ikke forsvinder fra sitet — de står jo på listen.
                    n = visningsnavn(afdeling, andelsklasse, navn)
                    if not n:
                        uden_cvr.append(f"ark {ark_årstal}: række uden både CVR og navn")
                        continue
                    fond_id = f"NAVN-{slug(n)}"
            elif rå_isin:
                problemer.append(f"Ark '{sheet.title}': ulæselig ISIN-værdi {rå_isin!r}")
                continue
            else:
                tomme_rækker += 1
                continue

            e = fonde.setdefault(
                fond_id,
                {"type": type_, "navne": [], "aar": set(), "lei": None,
                 "cvr": None, "hjemsted": None, "kontrolciffer_fejl": False},
            )
            vn = visningsnavn(afdeling, andelsklasse, navn)
            if vn and vn not in e["navne"]:
                e["navne"].append(vn)
            e["aar"] |= år
            e["lei"] = e["lei"] or c("lei")
            e["cvr"] = e["cvr"] or cvr
            e["hjemsted"] = e["hjemsted"] or c("hjemsted")
            if type_ == "isin" and fond_id in ugyldige_isin:
                e["kontrolciffer_fejl"] = True

            ark_år.setdefault(ark_årstal, set()).add(fond_id)
            antal += 1

        rap.append(f"Ark '{sheet.title}': {antal} datarækker læst.")

    if problemer and not fonde:
        rap.append("")
        rap.append("KRITISKE PROBLEMER:")
        rap.extend(f"  - {p}" for p in problemer)
        REPORT_FILE.write_text("\n".join(rap), encoding="utf-8")
        print("\n".join(rap))
        sys.exit("FEJL: parsning ufuldstændig — intet output skrevet.")
    if not fonde:
        sys.exit("FEJL: ingen fonde fundet — intet output skrevet.")

    uenige = sum(
        1 for årstal, ids in ark_år.items() for fid in ids if årstal not in fonde[fid]["aar"]
    )
    kun_2020 = sum(1 for e in fonde.values() if 2020 in e["aar"]) if 2020 not in ark_år else 0

    alle_år = sorted({y for e in fonde.values() for y in e["aar"]})
    aktuelt = max(ark_år)
    antal_pr_år = {y: sum(1 for e in fonde.values() if y in e["aar"]) for y in alle_år}
    med_isin = sum(1 for e in fonde.values() if e["type"] == "isin")
    uden_isin = len(fonde) - med_isin

    rap += [
        "",
        f"Fonde i alt: {len(fonde)}  (med ISIN: {med_isin}, uden ISIN: {uden_isin})",
        f"Aktuelt indkomstår (nyeste faneblad): {aktuelt}",
        "",
        "Registrerede fonde pr. indkomstår (kilde: kolonnen 'Registrerede år'):",
    ]
    for y in alle_år:
        note = "   [intet faneblad — kun fra årskolonnen]" if y not in ark_år else ""
        rap.append(f"  {y}: {antal_pr_år[y]}{note}")

    rap += [
        "",
        "Kontroller:",
        f"  Tomme rækker sprunget over: {tomme_rækker}",
        f"  Fonde hvor fanebladets år ikke stod i 'Registrerede år': {uenige}",
        f"  Fonde med 2020 i årskolonnen (2020 har intet faneblad): {kun_2020}",
        f"  ISIN med forkert kontrolciffer (beholdt, markeret): {len(ugyldige_isin)}"
        + (f" — {sorted(ugyldige_isin)}" if ugyldige_isin else ""),
        f"  ISIN-løse fonde uden både CVR og navn (udeladt): {len(uden_cvr)}"
        + (f" — {uden_cvr}" if uden_cvr else ""),
    ]
    if problemer:
        rap.append("  Advarsler (rækker sprunget over):")
        rap.extend(f"    - {p}" for p in problemer)

    flere_navne = [(k, v["navne"]) for k, v in fonde.items() if len(v["navne"]) > 1]
    rap.append(f"  Fonde registreret under flere navne: {len(flere_navne)}")
    for k, n in flere_navne[:3]:
        rap.append(f"    {k}: {n}")

    rap.append("")
    rap.append("Eksempler til efterprøvning:")
    for fid, e in list(fonde.items())[:3]:
        rap.append(f"  {fid}: {e['navne'][0] if e['navne'] else '(uden navn)'} — år: {sorted(e['aar'])}")
    for fid, e in fonde.items():
        if e["type"] == "uden_isin":
            rap.append(f"  {fid} (uden ISIN): {e['navne'][0]} — år: {sorted(e['aar'])}")
            break

    REPORT_FILE.write_text("\n".join(rap), encoding="utf-8")
    print("\n".join(rap))

    if verify_only:
        print("\n--verify: abis.json IKKE skrevet.")
        return

    out = {
        "genereret_utc": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "kilde": {
            "side": state["source_page"],
            "fil_url": state["file_url"],
            "filnavn": state["original_filename"],
            "sha256": state["sha256"],
            "offentliggjort_tekst": state.get("published_text"),
            "hentet_utc": state["fetched_at_utc"],
        },
        "seneste_indkomstaar": aktuelt,
        "alle_aar": alle_år,
        "antal_pr_aar": {str(y): n for y, n in antal_pr_år.items()},
        "antal_uden_isin": uden_isin,
        "fonde": {
            fid: {**e, "aar": sorted(e["aar"])}
            for fid, e in sorted(fonde.items())
        },
    }
    OUT_FILE.write_text(json.dumps(out, ensure_ascii=False, indent=1), encoding="utf-8")
    print(f"\nSkrev {OUT_FILE} med {len(fonde)} fonde.")


if __name__ == "__main__":
    main(verify_only="--verify" in sys.argv)
